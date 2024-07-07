from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill
import random
from task_data import task_data, task_order, days, letters


def get_names():
    global names
    global floors_dict
    floors_dict = {}
    with open("configuration_taches.txt") as f:
        text = f.readlines()
        print(text)
    f.close()

    # Get all names:
    names = []
    current_floor = 1
    for line in text:
        if line[0] == '-':
            current_floor += 1
            continue
        if line.strip() == '':
            continue
        names.append(line.strip())
        floors_dict[line.strip()] = current_floor
    random.shuffle(names)

def set_data():
    global task_scores_dict
    get_names()
    task_scores_dict = {}
    for name in names:
        task_scores_dict[name] = 0

def sort_names():
    """Sort names by their task scores, so the people with the least tasks go to the beginning of the list"""
    random.shuffle(names)
    names.sort(key=lambda x: task_scores_dict[x])
    print(names)
    print("Sorting names")

def has_any_task_this_day(name, day_index):
    day_row = day_index + 2
    for i in range(1, 13):
        cell_value = sheet[letters[i] + str(day_row)].value
        if sheet[letters[i] + str(day_row)].value == name:
            return True
        if cell_value is None:
            continue
        if "+" in cell_value:
            if name in cell_value.split("+"):
                return True
    return False

def already_has_task(name, task_name):
    task_cells = task_data[task_name]["cells"]
    for cell in task_cells:
        if str(sheet[cell].value) == name:
            return True
    return False

def get_day_index_from_cell(cell):
    return int(cell[-1]) - 2

def get_floor_residents(floor_index):
    temp_array = []
    for name in names:
        if floors_dict[name] == floor_index:
            temp_array.append(name)
    return temp_array

def get_lowest_fitting_person(task_name, cells_left):
    """Returns an array with [name, cell]"""
    sort_names()
    lowest_people = []
    lowest_value = task_scores_dict[names[0]]
    for name in names:
        if task_scores_dict[name] == lowest_value:
            lowest_people.append(name)
    random.shuffle(lowest_people)
    _cells = cells_left.copy()
    random.shuffle(_cells)

    for cell in _cells:
        for name in lowest_people:
            if has_any_task_this_day(name, get_day_index_from_cell(cell)):
                print(f"{name} already has a task this day, skipping..")
                continue

            if already_has_task(name, task_name):
                print(f"{name} already has this task, skipping..")
                continue
            return [name, cell]
    
    print("Failed to choose lowest fitting person")
    return [random.choice(lowest_people), random.choice(_cells)]

def get_task_by_cell(cell):
    for task in task_data.keys():
        if cell in task_data[task]["cells"]:
            return task
    return None

def read_existing_values():
    file_name = "preset_sheet.xlsx"
    _wb = load_workbook(file_name)
    _sheet = _wb["Sheet"]
    for letter_index in range(1, 13):
        for number in range(2, 9):
            formed_cell = letters[letter_index] + str(number)
            cell_value = _sheet[formed_cell].value
            _names = []
            if cell_value is not None:
                _names = [cell_value]
                print(cell_value)
                if "+" in cell_value:
                    _names = cell_value.split("+")

                if cell_value not in names:
                    print("/!\ wrong name")
                for n in _names:
                    give_task(n, get_task_by_cell(formed_cell), formed_cell, custom_color='BBBBBB')
                if "+" in cell_value:
                    sheet[formed_cell].value = cell_value
    sort_names()
                
            

def give_task(name, task_name, cell, custom_color=None):
    sheet[cell].value = name
    task_cost = task_data[task_name]["task_cost"]
    task_scores_dict[name] += task_cost
    color = get_task_cost_color(task_cost)
    if custom_color is not None:
        color = custom_color
    sheet[cell].fill = PatternFill('solid', fgColor = color)
    print(f"Gave {name} task {task_name} at {cell}")


def distribute_task(task_name):
    _cells = task_data[task_name]["cells"]
    _people_count_for_task = task_data[task_name]["people_per_grid"]
    _task_cost = task_data[task_name]["task_cost"]
    _task_limit = task_data[task_name]["task_limit"]
    _floor_specific = task_data[task_name]["floor_specific"]

    _floor_resident_list = []
    if _floor_specific is not None:
        _floor_resident_list = get_floor_residents(_floor_specific)

    for cell in _cells:
        if sheet[cell].value is not None:
            _cells.remove(cell)

    _task_count = 0
    for i in range(0, len(_cells)):

        if _floor_specific is not None:
            if len(_floor_resident_list) == 0:
                print("All floor residents have been assigned")
                return
        _people_to_add_to_cell = []

        for j in range(0, _people_count_for_task):
            temp = get_lowest_fitting_person(task_name, _cells)
            chosen_person = temp[0]
            if j == 0:
                chosen_cell = temp[1]
            if _floor_specific is not None:
                chosen_person = random.choice(_floor_resident_list)
                _floor_resident_list.remove(chosen_person)
            give_task(chosen_person, task_name, chosen_cell)
            _people_to_add_to_cell.append(chosen_person)

        _cells.remove(chosen_cell)

        if _people_count_for_task > 1:
            _str = "+".join(_people_to_add_to_cell)
            sheet[chosen_cell].value = _str

        _task_count += 1

        if _task_limit is not None:
            if _task_count >= _task_limit:
                print("Reached task limit")
                return

def adjust_cell_sizes():
    for i in range(1, 15):
        max_len = 0
        # For fun
        len_list = [len(val) for val in [sheet[letters[i]+str(j)].value for j in range(2, 15)] if val is not None]

        if len_list != []:
            max_len = max(len_list)
            if sheet.column_dimensions[letters[i]].width < max_len:
                sheet.column_dimensions[letters[i]].width = max_len

def get_task_cost_color(task_cost):
    #colors      0.5       1         2         3
    colors = ['a3ffb4', '4974a5', 'e23a08', 'a14242']
    current_color_index = 0
    if task_cost > 0.5:
        current_color_index = 1
    if task_cost > 1:
        current_color_index = 2
    if task_cost > 2:
        current_color_index = 3
    return colors[current_color_index]

def decorate_sheet():
    # Fill the task numbers at the top of the sheet
    for i in range(2, 13):
        if i == 2:
            sheet["B1"] = "1/2"
            continue
        sheet[letters[i-1]+"1"] = i


    # Fill the left side with days
    for i, day in enumerate(days):
        sheet["A" + str(i+2)] = day

    adjust_cell_sizes()
    sheet.merge_cells("C7:C8")


def print_debug_info():
    max_value = max(task_scores_dict, key=task_scores_dict.get)
    min_value = min(task_scores_dict, key=task_scores_dict.get)
    resident_count = len(task_scores_dict.keys())
    _sum = sum(task_scores_dict.values())
    mean = _sum / resident_count
    variance = 0
    value_list = []
    for name in names:
        variance = task_scores_dict[name] ** 2
        value_list.append(task_scores_dict[name])
    print(sorted(value_list))
    print(f"sum: {_sum}")
    print(f"resident count: {resident_count}")
    print(f"mean: {mean}")
    print(f"variance: {variance}")
    print(f"{max_value} {task_scores_dict[max_value]}")
    print(f"{min_value} {task_scores_dict[min_value]}")
    print(f"Valeure d'injustice: {task_scores_dict[max_value] - task_scores_dict[min_value]}")

if __name__ == "__main__":
    set_data()
    wb = Workbook()
    sheet = wb.active
    read_existing_values()
    for task in task_order:
        distribute_task(task)
    decorate_sheet()
    wb.save("sheet.xlsx")
    print_debug_info()